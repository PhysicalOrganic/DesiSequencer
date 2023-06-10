import pandas as pd
import matplotlib.pyplot as plt

from pathlib import Path
from openpyxl import Workbook

import numpy as np
import argparse

from utils import MonomerMasses
from plotting import plotTwoMassSpectra, plotSpots, plotMassSpectrum

from massCluster import findLargestMassPeak

def getArgs():
    pass

def CleanUpScan(
        scan: str, 
        mz_round: int = 2) -> dict:
    '''
    Takes the raw scan data formatted as a string and splits it
    into a dictionary of m/z:intensity pairs.

    Parameters
    ----------
    scan: str
        Original scan data formatted like:
        ...m/z, intensity;m/z, intensity;m/z, intensity;...

    mz_round: int
        Number of decimal points to round each m/z

    Returns
    ----------
        dict
    '''

    NewDict = {}

    scan = scan.split(';')
    for pair in scan:
        if len(pair) > 0:
            pair = pair.strip(' ')
            pair = pair.split(',')
            NewDict[round(float(pair[0]),2)] = round(float(pair[1]),1)

    return NewDict

def getAverageIntensity(
        scan: dict, 
        mass_thresh: float = 500) -> float:
    '''
    Gets the average intensity of all the masses of a scan
    which is formatted as a dictionary of mass:intensity key:value
    pairs. Returns the average intensity.

    scan: dict
        Dictionary of mass:intensity pairs which represents
        a mass spectrum scan

    mass_thresh: float
        Threshold overwhich masses are considered 
    '''
    return np.average([scan[x] for x in scan.keys() if x >= mass_thresh])

def findPeaks(
    scanData: pd.DataFrame, 
    parent_mass: float,
    intensity_threshold: int = 0.02,
    minimum_mass: int = 250,
    debug = True): #110000
    '''
    Identifies the signals associated with the loss of a monomer
    from the parent oligomer. The parent oligomer is identified 
    by the mass

    Parameters
    ----------
    ScanData: pd.DataFrame
        Dataframe containing two columns: Mass, Intensity. Typically the averaged
        mass spectrum of a particular sample (i.e., the average of multiple scans
        which are associated with a spot on a DESI plate)

    parent_mass: float
        Mass of the parent oligomer of length n which is used to find the 
        n - x oligomers resulting from sequencing. x represents the sequence
        iteration.

    intensity_threshold: float (default = 0.02)
        The percentage of intensity of the largest intensity signal
        within the scanData["Intensity"] column above which signals
        are considered for selection.

        Default is 2% of maximum intensity signal

    minimum_mass: int (default = )
        The mass below which signals are considered insignificant.

    Returns
    ----------
    monomers: dict
        Mass:intensity pairs of the selected "peaks" which will be used
        to subtract and thus get the mass loss. This mass loss will eventually 
        be matched to a monomer identity.
    '''

    # Convert intensity threshold to count
    intensity_threshold = intensity_threshold * scanData['Intensity'].max()

    # Get the largest mass in the spectrum
    LargestMassPeakIntensity = float(scanData[scanData['Mass'] == parent_mass].Intensity)

    # dictionary containing identity of monomers and the mass at which they are found
    monomers = {}

    # First "monomer" is not actually a monomer, it is the mass:intensity pair of the oligo which will lose a monomer once sequenced.
    monomers[parent_mass] = LargestMassPeakIntensity 

    # Get the lighest and heaviest monomers
    heaviest_monomer = max(MonomerMasses.keys())
    lightest_monomer = min(MonomerMasses.keys())

    current_peak = parent_mass
    while True:

        lower_bound = current_peak - heaviest_monomer
        upper_bound = current_peak - lightest_monomer

        tmp_df = scanData[(scanData['Mass'] >= lower_bound) & (scanData['Mass'] <= upper_bound)]
        #print(tmp_df.shape, ScanData.shape, 'lower bound',current_peak - heaviest_monomer, 'upper bound', current_peak - lightest_monomer)
        
        if debug:
            if len(monomers) % 2 == 0:
                color = 'purple'
            else:
                color = 'green'

            h = scanData[(scanData['Mass'] >= lower_bound) & (scanData['Mass'] <= upper_bound)].Intensity.max() * 1.05
            rect = plt.Rectangle((lower_bound, 0), width = upper_bound - lower_bound, height = h, color=color, alpha = 0.3)
            plt.gca().add_patch(rect)
        
        if tmp_df.empty:
            break

        # Get the new largest peak
        current_peak = float(tmp_df[tmp_df['Intensity'] == tmp_df['Intensity'].max()].Mass)
        monomers[current_peak] = float(tmp_df[tmp_df['Mass'] == current_peak].Intensity)

        if lower_bound <= minimum_mass:
            break

    if debug:
        print(f'Selected masses: {list(monomers.keys())}\n')

    return monomers

def MatchMasstoMonomer(
        MonomerMasses: dict, 
        SpectraMasses, 
        monomer_tolerance: float = 1, 
        endcap_mass: float = 262.084, 
        endcap_name: str = 'Tyr(OMe)', 
        debug: bool = False):
    '''
    Function which assigns a monomer identity to a given mass

    Monomer masses was a dictionary of mass:intensity pairs
    '''
    # MonomerMasses has the keys as the mass

    # Is there a reason Mass:intensity pairs are shipped together into this function?
    masses_identified_as_oligos = list(SpectraMasses.keys())

    # List of monomer names that were identified
    monomer_names = []

    for x in range(len(masses_identified_as_oligos)-1):

        # Test for endcap + last monomer
        # Difference in mass between the current oligomer
        # and the endcap (which would be the last monomer)
        endcap_difference = masses_identified_as_oligos[x] - endcap_mass
        # Differences between the endcap difference and each monomer.
        difs = {monomer_symbol:k - endcap_difference for k, monomer_symbol in MonomerMasses.items()}
        # Test if any of these are less than the tolerance
        for monomer_symbol, m in difs.items():
            if abs(m) <= monomer_tolerance:
                monomer_names.append(monomer_symbol)
                monomer_names.append(endcap_name)
                return monomer_names

        mass_loss = masses_identified_as_oligos[x] - masses_identified_as_oligos[x+1]
        
        # Iterate over all the possible monomers and their masses
        # Dictionary for storing the difference between the mass loss and a potential monomer.
        # The lowest of this dictionary's values is the most likely monomer
        differences = {}
        for mass, monomer_symbol in MonomerMasses.items():
            differences[monomer_symbol] = abs(round(mass - mass_loss, 2))

        if debug:
            from pprint import pprint
            print('Differences between mass loss and possible monomers')
            pprint(differences)
            print()
        
        lowest_match_difference = min(differences.values())
        match = list(differences.keys())[list(differences.values()).index(min(differences.values()))]

        if lowest_match_difference >= 1:
            print(f'WARNING: Mass difference for {masses_identified_as_oligos[x]} - {masses_identified_as_oligos[x+1]} = {round(mass_loss, 2)}')
            print(f'was larger than the acceptable tolerance for monomer identification\nTolerance: {monomer_tolerance}\tDiff: {lowest_match_difference}\n')

        monomer_names.append(match)

    return monomer_names

def sequenceFromDesiFile(
        file: Path, 
        debug: bool = False
        ):
    
    # Initiate plotting here so we can plot within other functions
    fig, ax = plt.subplots(1,1, figsize = (12,6))

    name = str(file.stem)

    if debug:
        print(f"Processing File: {name}")

    with open(file, "r") as f:
        RAWdata = f.read()

    # split combined scans into individual scan events
    RAWdata = RAWdata.split("|")

    #### CLEANING UP RAW FILE DATA ####
    # cleaning up scans to remove brackets and commas and whatnot and format them into a dictionary
    CombinedScans = []   
    for Scan in RAWdata:
        NewScan = CleanUpScan(Scan)
        if (len(NewScan) != 0):
            CombinedScans.append(NewScan)

    if debug:
        print("Number of Scans (time points):", len(CombinedScans))

    AverageScanIntensities = []

    # plot the average intensity of each scan to visualize spots
    for i in CombinedScans:
        AverageScanIntensities.append(getAverageIntensity(i, mass_thresh=500))

    #PlotSpots(len(CombinedScans), AverageScanIntensity)

    # identify the spots based on intensity
    IsolateSpots = []
    TempSpots = []

    # this is the scan intensity that you check to see what is noise and what is oligos
    IntensityThreshold = 20000

    # this is how many scans should make up one spot to get rid of the spikes there sometimes are
    SpotWidth = 6

    count = 0
    for i, intensity in enumerate(AverageScanIntensities): # Go through all avg intensities
        if intensity > IntensityThreshold and count < 9: # If it is above the threshold and #TODO less than the 9th scan intensity?
            count += 1
            TempSpots.append(i) # Add it to the temporary spots
        elif count >= SpotWidth: # If the count is above or equal to the spot width (6 currently)
            IsolateSpots.append(TempSpots) # Add this to the isolate spots
            TempSpots = [] # Reset the temp spots?
            count = 0

    print(f'Number of spots Found: {len(IsolateSpots)}\n')
    print(f"Spots found at scan Numbers: {IsolateSpots}\tNumber of scans: {len(CombinedScans)}\n")    
    
    # List of dictionaries which have mass:intensity pairs
    # Is the length of the number of spots
    CombinedSpotScans = []

    # Iterate through spot scans
    for i in IsolateSpots: # For every list of scans (dictionaries of mass:intensity pairs), these are combined because they represent a single sample
        BasePeak = CombinedScans[i[0]] # This gets the mass scan of the first peak (i[0] is the first index of the peak)
        for j in range (1, len(i)): # Each peak has multiple indices over which it exists. Iterate over them
            for k in CombinedScans[j].keys(): # Get the other 
                if k in BasePeak.keys():
                    BasePeak[k] += CombinedScans[j][k] # I think we are adding the intensity of those masses from each scan to a single 
                else:
                    BasePeak[k] = CombinedScans[j][k]
        CombinedSpotScans.append(BasePeak)

    columnNum = 1
    SpotCount = 1

    # Make a workbook to save the data
    wb = Workbook()
    ws = wb.active

    # Iterate over every spot (which is the combined intensities over all scans which were done for that spot)
    for x in CombinedSpotScans: 

        ScansTemp = pd.DataFrame(list(x.items()), columns=['Mass', 'Intensity'])

        # Find the parent oligomer mass
        parent_mass = findLargestMassPeak(scanData=ScansTemp, debug=debug)
        parent_mass_intensity = float(ScansTemp[ScansTemp["Mass"] == parent_mass].Intensity)

        if debug:
            print(f'Largest mass in spectrum: {parent_mass}\tIntensity: {parent_mass_intensity}\n')

        # Find the peaks
        FoundSpectraMassesIntensity = findPeaks(scanData = ScansTemp, parent_mass = parent_mass, debug=debug)

        # Convert the mass:intensity pairs into list of masses
        FoundSpectraMasses = list(FoundSpectraMassesIntensity.keys())

        # Match the list of masses to a particular monomer loss
        MassMatches = MatchMasstoMonomer(MonomerMasses, FoundSpectraMassesIntensity, endcap_mass=endcap_mass, endcap_name=endcap_name)

        # Iterate over the masses
        for i, mass_found_in_spectrum in enumerate(FoundSpectraMasses):
            # If 
            if i < len(MassMatches):
                if MassMatches[i] == endcap_name:
                    print(endcap_mass, endcap_name, "(Endcap)")
                    ws.cell(row=i+1, column=columnNum).value = round(endcap_mass, 2)
                    ws.cell(row=i+1, column=columnNum+1).value = endcap_name
                else:
                    print(mass_found_in_spectrum, MassMatches[i])
                    ws.cell(row=i+1, column=columnNum).value = mass_found_in_spectrum
                    ws.cell(row=i+1, column=columnNum+1).value = MassMatches[i]
            else:
                print("ERROR I DONT UNDERSTAND", MassMatches, i)

        # Increment the column number by 3 for the next spot
        columnNum +=3

        x, y = zip(*sorted(x.items()))  
        ax.stem(x, y, linefmt='black', markerfmt='')
        ax.set_title(f"Spot {SpotCount} of {len(IsolateSpots)} ({file.stem})")
        ax.set_ylabel('Counts')
        ax.set_xlabel(r'$m/z$')
        #FoundSpectraMassesIntensity = {k:v for k,v in FoundSpectraMassesIntensity.items() if v > 100000}
        ax.stem(FoundSpectraMassesIntensity.keys(), FoundSpectraMassesIntensity.values(), linefmt='red', markerfmt='') #, width=1.5

        if debug:
            ax.text(parent_mass*1.005, parent_mass_intensity * 1.005, 'Parent peak', color='red')
        plt.show()

        SpotCount += 1
        savePath = Path(file.parent / str(name + ".xlsx"))
        print(f'\nFile saved at: {savePath.absolute()}')
        wb.save(savePath)   
        exit()

if __name__ == "__main__":
    # Define a file to 
    f = Path('./DesiSequencer/data/Oligomers March 2023/manualProfiling_mayaAngelou_spot1.txt')
    
    # Set some parameters
    endcap_mass=262.084
    endcap_name='Tyr(OMe)'
    debug = True

    sequenceFromDesiFile(f, debug=debug)


    
